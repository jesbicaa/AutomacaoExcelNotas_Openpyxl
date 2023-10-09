# pip install selenium
# pip install webdriver_manager
# pip install pywin32
# pip install openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from time import sleep
import openpyxl
import smtplib
import re

class Scrappy:
    def iniciar(self):
        self.email_usuario()
        self.login_site()
        self.raspagem_de_dados()
        self.colocar_dados_excel()
        self.enviar_email_usuario()
    def email_usuario(self):
        self.email = input('Digite o email para receber o relatorio de média:\n')
        self.email.lower()

        padrao = re.search(r'[a-zA-Z0-9_-]+@[a-zA-Z0-9]+\.[a-zA-Z]{1,3}$', self.email)

        if padrao:
            print('email Valido')

        else:
            print('Digite um email valido!!!')
            self.email_usuario()

    def login_site(self):
        self.prontuario = input('Digite seu prontuario: ')
        self.senha = input('Digite sua senha do Suap: ')

        servico = Service(ChromeDriverManager().install())
        self.navegador = webdriver.Chrome(service=servico)

        # acessar o site
        self.link = 'https://suap.ifsp.edu.br/accounts/login/?next=/'
        self.navegador.get(self.link)
        sleep(1)

        # logar em uma conta
        self.navegador.find_element('xpath', '//*[@id="id_username"]').send_keys(self.prontuario)
        sleep(2)
        self.navegador.find_element('xpath', '//*[@id="id_password"]').send_keys(self.senha)
        sleep(2)
        self.navegador.find_element('xpath', '/html/body/div[1]/main/div[1]/form/div[5]/input').click()
        sleep(3)

    def raspagem_de_dados(self):
        self.link = 'https://suap.ifsp.edu.br/edu/aluno/' + self.prontuario.upper() + '/?tab=boletim'
        self.lista_notas = []
        self.navegador.get(self.link)
        sleep(2)
        materia = 1
        for p in range(14):
            bimestre = 10
            for i in range(4):
                lista_bim = self.navegador.find_element('xpath',
                    f'/html/body/div[1]/main/div[4]/div[2]/div[9]/div/div[2]/div/div/table[1]/tbody/tr[{materia}]/td[{bimestre}]')
                if lista_bim.text == '-':
                    lista_bim = 0
                    self.lista_notas.append(float(lista_bim))
                    sleep(1)
                    bimestre += 2
                else:
                    self.lista_notas.append(float(lista_bim.text))
                    sleep(1)
                    bimestre += 2
            materia += 1
            print(str(p) + ' - Proxima Materia')
        self.navegador.quit()
        print('Raspagem de notas executada com sucesso')

    def colocar_dados_excel(self):
        # importar a tabela
        tabela = openpyxl.load_workbook('Fiquei_na_media-Caires.xlsx')
        tabela_de_media = tabela.active

        n1 = 0
        n2 = 1
        n3 = 2
        n4 = 3
        index = 4
        for pp in range(14):
            tabela_de_media.cell(column=3, row=index, value=self.lista_notas[n1])
            tabela_de_media.cell(column=4, row=index, value=self.lista_notas[n2])
            tabela_de_media.cell(column=5, row=index, value=self.lista_notas[n3])
            tabela_de_media.cell(column=6, row=index, value=self.lista_notas[n4])
            n1 += 4
            n2 += 4
            n3 += 4
            n4 += 4
            index += 1
        tabela.save('Fiquei_na_media-Caires.xlsx')
        print('Planilha atualizada com sucesso!')

    def enviar_email_usuario(self):
        endereco = input('Digite seu email aqui: ')
        senha = input('Digite sua senha aqui: ')
        arquivo = 'Fiquei_na_media-Caires.xlsx'

        try:
            servidor_email = smtplib.SMTP('smtp.gmail.com', 587)
            servidor_email.starttls()
            servidor_email.login(endereco, senha)

            remetente = 'seu_email@gmail.com'
            destinatarios = self.email
            assunto = 'Excel para Saber se Ficou na Média'
            conteudo = ('<p>Segue o Excel para saber se ficou na média.</p>'
                        '<p>att.,</p>'
                        '<p>Jesbica</p>')


            message = (f"From:{remetente}\nSubject:{assunto}\n\n{conteudo}\n"
                       f"{arquivo}, maintype='application',subtype='octet-stream', filename='Fiquei_na_media-Caires.xlsx'")

            servidor_email.sendmail(remetente, destinatarios, message)
        except Exception as e:
            print(f"Erro ao enviar email: {e}")
        finally:
            servidor_email.quit()

start = Scrappy()
start.iniciar()
